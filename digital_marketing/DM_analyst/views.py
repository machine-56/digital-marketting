import openpyxl
from django.shortcuts import redirect, render
from django.contrib import messages
from accounts.models import CustomUser, LeaveRequest, Attendance
from dm_excutive.models import Lead, LeadAssignment

from django.utils.crypto import get_random_string
from django.core.mail import send_mail
from django.conf import settings

from accounts.models import Attendance
from datetime import date

# Create your views here.
def dma_home(request):
    return render(request, 'DM_analyst/dma_home.html')


def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        skipped_entries = []

        for row in sheet.iter_rows(min_row=2, values_only=True):  
            if not row or not row[0]:
                continue

            name, email, phone, address, gender = row

            if ' ' in name:
                first_name, last_name = name.split(' ', 1)
            else:
                first_name, last_name = name, ''

            email_exists = CustomUser.objects.filter(email=email).exists()
            phone_exists = CustomUser.objects.filter(phone=phone).exists()
            name_exists = CustomUser.objects.filter(first_name=first_name, last_name=last_name).exists()

            if email_exists or phone_exists or name_exists:
                reason = []
                if email_exists:
                    reason.append("email")
                if phone_exists:
                    reason.append("phone")
                if name_exists:
                    reason.append("name")

                skipped_entries.append(f"{name} skipped (duplicate {', '.join(reason)})")
                continue

            valid_genders = ['male', 'female', 'other']
            gender_value = gender.lower() if gender and gender.lower() in valid_genders else None

            username = f"{first_name.lower()}_{last_name.lower()}" if last_name else first_name.lower()
            base_username = username
            counter = 1
            while CustomUser.objects.filter(username=username).exists():
                username = f"{base_username}{counter}"
                counter += 1

            password = get_random_string(length=12)

            user = CustomUser.objects.create(
                first_name=first_name,
                last_name=last_name,
                username=username,
                email=email,
                phone=phone,
                address=address,
                role='dm_executive',
                gender=gender_value,
                company_name=request.user.company_name,
                profile_image='profile_images/image-1.png',
                is_approved=True,
                status=1
            )
            user.set_password(password)
            user.save()

            subject = "Welcome to the Digital Marketing Team – Your Account Details Inside"
            message = (
                f"Hi {first_name},\n\n"
                f"We are excited to welcome you to the Digital Marketing team at {request.user.company_name}!\n\n"
                f"Your account has been successfully created. Below are your login credentials:\n\n"
                f"-------------------------------------------\n"
                f"Username: {username}\n"
                f"Password: {password}\n"
                f"-------------------------------------------\n\n"
                f"Please log in to your account using the credentials above and reset your password immediately for security purposes.\n\n"
                f"Login URL: http://127.0.0.1:8000/login/  (update this link accordingly)\n\n"
                f"If you encounter any issues, feel free to contact our support team.\n\n"
                f"Warm regards,\n"
                f"Digital Marketing Team\n"
                f"{request.user.company_name}"
            )

            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email])

        messages.success(request, 'Excel data uploaded successfully!')
        if skipped_entries:
            for entry in skipped_entries:
                messages.warning(request, entry)

        return redirect('upload_excel')

    return render(request, 'DM_analyst/upload_excel.html')


def view_executives(request):
    executives = CustomUser.objects.filter(role='dm_executive', is_approved=True)
    return render(request, 'DM_analyst/view_executives.html', {'executives': executives})


def assign_leads(request, executive_id):
    executive = CustomUser.objects.get(id=executive_id, role='dm_executive')

    if request.method == 'POST':
        selected_lead_ids = request.POST.getlist('lead_checkbox')
        leads = Lead.objects.filter(id__in=selected_lead_ids)

        for lead in leads:
            lead.executive = executive
            lead.is_assigned = True
            lead.save()

            LeadAssignment.objects.create(lead=lead, executive=executive)

        messages.success(request, f"Leads successfully assigned to {executive.first_name} {executive.last_name}.")
        return redirect('view_executives')

    unassigned_leads = Lead.objects.filter(is_assigned=False, is_verified = True)
    return render(request, 'DM_analyst/assign_leads.html', {
        'executive': executive,
        'leads': unassigned_leads
    })


def upload_leads(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        skipped_entries = []

        for row in sheet.iter_rows(min_row=2, values_only=True):  
            if not row or not row[0]:
                continue

            lead_name, email, phone, platform = row

            email_exists = Lead.objects.filter(email=email).exists()
            phone_exists = Lead.objects.filter(phone=phone).exists()

            if email_exists or phone_exists:
                reason = []
                if email_exists:
                    reason.append("email")
                if phone_exists:
                    reason.append("phone")
                skipped_entries.append(f"{lead_name} skipped (duplicate {', '.join(reason)})")
                continue

            lead = Lead(
                lead_name=lead_name,
                email=email,
                phone=phone,
                platform=platform.lower(),
                status='pending'
            )
            lead.save()

        messages.success(request, 'Lead data uploaded successfully!')
        if skipped_entries:
            for entry in skipped_entries:
                messages.warning(request, entry)

        return redirect('upload_leads')

    return render(request, 'DM_analyst/upload_leads.html')


def dma_view_leads(request):
    filter_option = request.GET.get('filter', 'all')

    if filter_option == 'verified':
        leads = Lead.objects.filter(is_verified=True)
    elif filter_option == 'pending':
        leads = Lead.objects.filter(is_verified=False)
    else:
        leads = Lead.objects.all()

    if request.method == 'POST':
        selected_ids = request.POST.getlist('lead_checkbox')
        Lead.objects.filter(id__in=selected_ids).update(status='verified', is_verified=True)
        messages.success(request, 'Selected leads have been verified.')
        return redirect('dma_view_leads')

    leads = Lead.objects.all()
    return render(request, 'DM_analyst/view_leads.html', {'leads': leads, 'filter': filter_option})


def dma_mark_attendance_and_apply_leave(request):
    today = date.today()

    leaves_today = LeaveRequest.objects.filter(
        user=request.user, 
        from_date__lte=today, 
        to_date__gte=today,
        status='approved'
    )

    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'attendance':
            status = request.POST.get('status')

            leave_conflict = LeaveRequest.objects.filter(
                user=request.user, 
                from_date__lte=today, 
                to_date__gte=today, 
                status='approved'
            )

            if leave_conflict.exists():
                messages.error(request, 'You cannot mark attendance during approved leave.')
            elif Attendance.objects.filter(user=request.user, date=today).exists():
                messages.warning(request, 'Attendance already marked for today.')
            else:
                Attendance.objects.create(user=request.user, date=today, status=status)
                messages.success(request, f'Attendance marked as {status.title()} for today.')

        elif form_type == 'leave':
            from_date = request.POST.get('from_date')
            to_date = request.POST.get('to_date')
            leave_type = request.POST.get('leave_type')
            reason = request.POST.get('reason')

            LeaveRequest.objects.create(
                user=request.user,
                from_date=from_date,
                to_date=to_date,
                leave_type=leave_type,
                reason=reason
            )
            messages.success(request, 'Leave request submitted successfully.')

        return redirect('mark_attendance_and_apply_leave')

    return render(request, 'DM_analyst/attendance.html', {
        'today': today,
        'leaves_today': leaves_today
    })


def change_password(request):
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if not request.user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')
            return redirect('change_password')

        if new_password != confirm_password:
            messages.error(request, 'New password and confirmation do not match.')
            return redirect('change_password')

        if (len(new_password) < 8 or
            not any(char.isupper() for char in new_password) or
            not any(char.isdigit() for char in new_password) or
            not any(char in '!@#$%^&*()_+-=[]{}|;:,.<>?/~' for char in new_password)):
            messages.error(request, 'Password must be at least 8 characters long and contain at least one uppercase letter, one digit, and one special character.')
            return redirect('change_password')
        
        request.user.set_password(new_password)
        request.user.save()
        messages.success(request, 'Password changed successfully. Please log in again.')
        return redirect('login_fn')

    return render(request, 'DM_analyst/change_password.html')
